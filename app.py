import csv
import os
import sqlite3
import tempfile
import webbrowser
from pathlib import Path
from datetime import date, datetime

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import sys

def resource_path(relative_path: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    return base / relative_path


from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.utils import simpleSplit
from babel.dates import format_date

# =========================
# Config
# =========================
APP_DIR = Path(__file__).parent
DATA_DIR = Path.home() / "Documents" / "Certificados_Escoteiros"
DATA_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = DATA_DIR / "membros.db"

# Template do certificado (imagem já cortada/ajustada)
TEMPLATE_PNG = resource_path("Certificado_limpo.png")

# Logo do grupo (coloque um PNG com esse nome na pasta do projeto)
LOGO_PNG = resource_path("logo_grupo.png")


PAGE_W, PAGE_H = A4

CHECK_OFF = "☐"
CHECK_ON = "☑"

SECAO_ORDER = [
    "Lobinhos",
    "Escoteiros",
    "Sênior",
    "Pioneiros",
    "Chefes",
    "Diretoria",
    "Outro",
]
SECAO_RANK = {s: i for i, s in enumerate(SECAO_ORDER)}

HIST_ATIVIDADES_PATH = APP_DIR / "historico_atividades.txt"
MAX_HIST = 25


# =========================
# Helpers
# =========================
def slug(text: str) -> str:
    """Deixa seguro para nome de arquivo no Windows."""
    invalid = '<>:"/\\|?*'
    for ch in invalid:
        text = text.replace(ch, "")
    text = text.strip().replace(" ", "_")
    return text[:60] if text else "atividade"


def parse_data(data_txt: str) -> date:
    data_txt = (data_txt or "").strip()
    if not data_txt:
        raise ValueError("Data vazia")

    if "/" in data_txt:
        d, m, a = data_txt.split("/")
        return date(int(a), int(m), int(d))
    # ISO
    return date.fromisoformat(data_txt)


def load_hist_atividades() -> list[str]:
    if not HIST_ATIVIDADES_PATH.exists():
        return []
    lines = [l.strip() for l in HIST_ATIVIDADES_PATH.read_text(encoding="utf-8").splitlines()]
    # remove vazios e duplicados preservando ordem
    out = []
    seen = set()
    for x in lines:
        if x and x.lower() not in seen:
            out.append(x)
            seen.add(x.lower())
    return out[:MAX_HIST]


def save_hist_atividades(nova: str):
    nova = (nova or "").strip()
    if not nova:
        return
    atual = load_hist_atividades()
    # coloca no topo
    atual = [x for x in atual if x.lower() != nova.lower()]
    atual.insert(0, nova)
    atual = atual[:MAX_HIST]
    HIST_ATIVIDADES_PATH.write_text("\n".join(atual), encoding="utf-8")


# =========================
# Banco (SQLite)
# =========================
def db_init():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS membros(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                secao TEXT NOT NULL
            )
            """
        )
        conn.commit()


def db_add(nome, secao):
    nome = (nome or "").strip()
    secao = (secao or "").strip() or "Outro"
    if not nome:
        return
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            "INSERT INTO membros(nome, secao) VALUES(?, ?)",
            (nome, secao),
        )
        conn.commit()


def db_list():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.execute("SELECT id, nome, secao FROM membros")
        rows = cur.fetchall()

    # ordena por secao (ordem fixa) e nome
    def key(r):
        _id, nome, secao = r
        return (SECAO_RANK.get(secao, 999), (nome or "").lower())

    return sorted(rows, key=key)


def db_delete(mid):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("DELETE FROM membros WHERE id=?", (mid,))
        conn.commit()


def db_find_by_name(nome: str):
    """Busca por nome exato (case-insensitive). Retorna (id, nome, secao) ou None."""
    nome = (nome or "").strip()
    if not nome:
        return None
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.execute(
            "SELECT id, nome, secao FROM membros WHERE lower(nome)=lower(?) LIMIT 1",
            (nome,),
        )
        return cur.fetchone()


# =========================
# Texto auto-ajustável
# =========================
def draw_left_fit(c, text, x_left, y, max_width, font="Helvetica-Bold", font_size=13, min_size=9):
    size = font_size
    while size >= min_size:
        c.setFont(font, size)
        if c.stringWidth(text, font, size) <= max_width:
            c.drawString(x_left, y, text)
            return
        size -= 1
    c.setFont(font, min_size)
    c.drawString(x_left, y, text)


def draw_center_wrap(c, text, cx, y_top, max_width, line_height,
                     font="Helvetica", font_size=11, max_lines=2):
    lines = simpleSplit(text, font, font_size, max_width)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        lines[-1] = lines[-1] + "..."
    y = y_top
    c.setFont(font, font_size)
    for line in lines:
        c.drawCentredString(cx, y, line)
        y -= line_height


# =========================
# PDF (3 por A4)
# =========================
def draw_cert(c, x, y, w, h, nome, atividade, local, dias, linha_evento, cidade, dt):
    # Fundo do certificado
    c.drawImage(
        str(TEMPLATE_PNG),
        x,
        y,
        width=w,
        height=h,
        preserveAspectRatio=True,
        mask="auto",
    )

    cx = x + w / 2
    max_w = w * 0.75

    # Linha: "Certificamos que" + Nome (mesma linha)
    linha_y = y + h * 0.56
    c.setFont("Helvetica", 11)
    c.drawString(x + w * 0.22, linha_y, "Certificamos que")

    nome_x = x + w * 0.42
    nome_max_w = (x + w * 0.78) - nome_x
    draw_left_fit(c, nome, nome_x, linha_y, nome_max_w, "Helvetica-Bold", 13, 9)

    # Texto principal (atividade/local/dias)
    draw_center_wrap(
        c,
        f"Participou do {atividade}, realizada em",
        cx,
        y + h * 0.49,
        max_w,
        13,
        font="Helvetica",
        font_size=11,
        max_lines=2,
    )

    draw_center_wrap(
        c,
        f"{local} nos dias {dias}.",
        cx,
        y + h * 0.44,
        max_w,
        13,
        font="Helvetica",
        font_size=11,
        max_lines=2,
    )

    # Linha do evento (negrito)
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(cx, y + h * 0.34, linha_evento)

    # Cidade + data por extenso
    data_extenso = format_date(dt, format="d 'de' MMMM 'de' y", locale="pt_BR")
    c.setFont("Helvetica", 11)
    c.drawRightString(x + w * 0.88, y + h * 0.25, f"{cidade}, {data_extenso}")


def gerar_pdf(participantes, atividade, local, dias, linha_evento, cidade, dt, output_file: str):
    if not TEMPLATE_PNG.exists():
        raise FileNotFoundError(f"Não encontrei '{TEMPLATE_PNG.name}' na pasta do app.")

    out = Path(output_file)

    c = canvas.Canvas(str(out), pagesize=A4)

    margin_x = 10 * mm
    margin_y = 10 * mm
    slot_w = PAGE_W - 2 * margin_x
    slot_h = (PAGE_H - 2 * margin_y) / 3

    i = 0
    while i < len(participantes):
        for pos in range(3):
            if i >= len(participantes):
                break
            x = margin_x
            y = PAGE_H - margin_y - (pos + 1) * slot_h

            draw_cert(
                c,
                x,
                y,
                slot_w,
                slot_h,
                nome=participantes[i],
                atividade=atividade,
                local=local,
                dias=dias,
                linha_evento=linha_evento,
                cidade=cidade,
                dt=dt,
            )
            i += 1
        c.showPage()

    c.save()
    return out


def gerar_preview_pdf(nome, atividade, local, dias, linha_evento, cidade, dt):
    """Gera PDF temporário com 1 certificado e abre."""
    if not TEMPLATE_PNG.exists():
        raise FileNotFoundError(f"Não encontrei '{TEMPLATE_PNG.name}' na pasta do app.")

    tmp_dir = Path(tempfile.gettempdir())
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = tmp_dir / f"preview_cert_{stamp}.pdf"

    c = canvas.Canvas(str(out), pagesize=A4)

    # usa o primeiro slot (topo)
    margin_x = 10 * mm
    margin_y = 10 * mm
    slot_w = PAGE_W - 2 * margin_x
    slot_h = (PAGE_H - 2 * margin_y) / 3
    x = margin_x
    y = PAGE_H - margin_y - 1 * slot_h

    draw_cert(c, x, y, slot_w, slot_h, nome, atividade, local, dias, linha_evento, cidade, dt)
    c.showPage()
    c.save()

    webbrowser.open(out.as_uri())


# =========================
# Import/Export (CSV/Excel)
# =========================
def read_rows_from_file(path: str):
    p = Path(path)
    ext = p.suffix.lower()

    if ext == ".csv":
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return list(reader)

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except Exception:
            raise RuntimeError("Para importar Excel (.xlsx) instale: pip install openpyxl")

        wb = openpyxl.load_workbook(p)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for i, v in enumerate(r):
                if i < len(headers):
                    d[headers[i]] = "" if v is None else str(v).strip()
            rows.append(d)
        return rows

    raise RuntimeError("Arquivo inválido. Use .csv ou .xlsx")


def normalize_cols(d: dict):
    # normaliza: lower + remove espaços extras
    m = { (k or "").strip().lower(): (str(v).strip() if v is not None else "") for k, v in d.items() }

    # aceita vários nomes possíveis de coluna
    nome = (
        m.get("nome")
        or m.get("name")
        or m.get("membro")
        or m.get("participante")
        or m.get("nome associados")     # <- seu Excel
        or m.get("nome associado")
        or m.get("associado")
        or ""
    )

    secao = (
        m.get("secao")
        or m.get("seção")
        or m.get("section")
        or m.get("associação")          # <- seu Excel
        or m.get("associacao")
        or ""
    )

    return nome.strip(), secao.strip()


def write_csv(path: str, rows: list[dict]):
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else ["nome", "secao"])
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


# =========================
# UI (Tkinter)
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Certificados Escoteiros")
        self.geometry("1050x680")
        self.minsize(1050, 680)

        db_init()

        left = ttk.Frame(self, padding=10)
        right = ttk.Frame(self, padding=10)
        left.pack(side="left", fill="both", expand=True)
        right.pack(side="right", fill="both", expand=True)

        # ---------- Cadastro ----------
        ttk.Label(left, text="Cadastro de membro", font=("Segoe UI", 12, "bold")).pack(anchor="w")

        self.nome_var = tk.StringVar()
        self.secao_var = tk.StringVar(value="Escoteiros")

        form = ttk.Frame(left)
        form.pack(fill="x", pady=6)

        ttk.Label(form, text="Nome:").grid(row=0, column=0, sticky="w")
        ttk.Entry(form, textvariable=self.nome_var, width=40).grid(row=0, column=1, sticky="w", padx=6)

        ttk.Label(form, text="Seção:").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Combobox(
            form,
            textvariable=self.secao_var,
            width=37,
            state="readonly",
            values=SECAO_ORDER,
        ).grid(row=1, column=1, sticky="w", padx=6)

        ttk.Button(left, text="Adicionar", command=self.add_membro).pack(anchor="w", pady=6)
        ttk.Separator(left).pack(fill="x", pady=10)

        # ---------- Busca + Lista ----------
        ttk.Label(left, text="Membros (clique no ☑ para marcar)", font=("Segoe UI", 11, "bold")).pack(anchor="w")

        self.busca_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.busca_var).pack(fill="x", pady=4)
        self.busca_var.trace_add("write", lambda *_: self.refresh_list())

        cols = ("sel", "nome", "secao")
        self.tree = ttk.Treeview(left, columns=cols, show="headings", height=18)
        self.tree.heading("sel", text="")
        self.tree.heading("nome", text="Nome")
        self.tree.heading("secao", text="Seção")

        self.tree.column("sel", width=40, anchor="center")
        self.tree.column("nome", width=330, anchor="w")
        self.tree.column("secao", width=120, anchor="w")
        self.tree.pack(fill="both", expand=True)

        self.tree.bind("<Button-1>", self.on_tree_click)
        self.tree.bind("<Double-1>", self.remover_tree_item)

        btns = ttk.Frame(left)
        btns.pack(fill="x", pady=6)
        ttk.Button(btns, text="Marcar todos", command=self.marcar_todos).pack(side="left")
        ttk.Button(btns, text="Desmarcar todos", command=self.desmarcar_todos).pack(side="left", padx=6)
        ttk.Button(btns, text="Importar (CSV/XLSX)", command=self.importar_membros).pack(side="left", padx=6)
        ttk.Button(btns, text="Exportar marcados", command=self.exportar_marcados).pack(side="left")

        self.itens = {}  # iid(mid) -> {"checked": bool, "nome": str, "secao": str}

        # ---------- Geração ----------
        ttk.Label(right, text="Gerar certificados", font=("Segoe UI", 12, "bold")).pack(anchor="w")

        # Atividade com histórico
        self.atividade_var = tk.StringVar()
        self.local_var = tk.StringVar()
        self.dias_var = tk.StringVar()
        self.cidade_var = tk.StringVar(value="Poços de Caldas")
        self.linha_evento_var = tk.StringVar(value="03/03/1991 – 34 anos da fundação do 103/MG – GEPIN")
        hoje = date.today()
        self.data_var = tk.StringVar(value=hoje.strftime("%d/%m/%Y"))

        def row_entry(label, var):
            f = ttk.Frame(right)
            f.pack(fill="x", pady=5)
            ttk.Label(f, text=label, width=18).pack(side="left")
            ttk.Entry(f, textvariable=var).pack(side="left", fill="x", expand=True)

        def row_combo(label, var, values):
            f = ttk.Frame(right)
            f.pack(fill="x", pady=5)
            ttk.Label(f, text=label, width=18).pack(side="left")
            cb = ttk.Combobox(f, textvariable=var, values=values, state="normal")
            cb.pack(side="left", fill="x", expand=True)
            return cb

        self.cb_atividade = row_combo("Atividade:", self.atividade_var, load_hist_atividades())
        row_entry("Local:", self.local_var)
        row_entry("Dias:", self.dias_var)
        row_entry("Cidade:", self.cidade_var)
        row_entry("Linha evento:", self.linha_evento_var)
        row_entry("Data (DD/MM/AAAA):", self.data_var)

        ttk.Separator(right).pack(fill="x", pady=10)

        # Botões
        actions = ttk.Frame(right)
        actions.pack(fill="x", pady=6)

        ttk.Button(actions, text="Pré-visualizar (1)", command=self.preview).pack(side="left")
        ttk.Button(actions, text="Gerar PDF (3 por folha A4)", command=self.gerar).pack(side="left", padx=8)

        # Logo abaixo
        self.logo_label = ttk.Label(right)
        self.logo_label.pack(anchor="w", pady=10)
        self.load_logo()

        ttk.Label(
            right,
            text="Dica: duplo clique em um membro remove do cadastro.",
            foreground="#555",
        ).pack(anchor="w", pady=4)

        self.refresh_list()

    # ---------- Logo ----------
    def load_logo(self):
        try:
            if LOGO_PNG.exists():
                # PhotoImage nativo (PNG ok)
                img = tk.PhotoImage(file=str(LOGO_PNG))
                # reduz se estiver grande (ex: 600px)
                w, h = img.width(), img.height()
                max_w = 280
                if w > max_w and w > 0:
                    factor = int(w / max_w) + 1
                    img = img.subsample(factor, factor)
                self._logo_ref = img
                self.logo_label.configure(image=img, text="")
            else:
                self.logo_label.configure(text="(Coloque 'logo_grupo.png' na pasta do projeto para aparecer aqui.)")
        except Exception:
            self.logo_label.configure(text="Não consegui carregar a logo (use PNG).")

    # ---------- Lista ----------
    def refresh_list(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.itens.clear()

        termo = self.busca_var.get().strip().lower()

        for mid, nome, secao in db_list():
            label = f"{nome} {secao}".lower()
            if termo and termo not in label:
                continue

            iid = str(mid)
            self.tree.insert("", "end", iid=iid, values=(CHECK_OFF, nome, secao))
            self.itens[iid] = {"checked": False, "nome": nome, "secao": secao}

    def on_tree_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not row_id:
            return
        if col == "#1":
            current = self.itens[row_id]["checked"]
            self.itens[row_id]["checked"] = not current
            self.tree.set(row_id, "sel", CHECK_ON if not current else CHECK_OFF)

    def remover_tree_item(self, event):
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return
        nome = self.itens[row_id]["nome"]
        if not messagebox.askyesno("Remover", f"Remover '{nome}' do cadastro?"):
            return
        db_delete(int(row_id))
        self.refresh_list()

    def marcar_todos(self):
        for iid in self.itens:
            self.itens[iid]["checked"] = True
            self.tree.set(iid, "sel", CHECK_ON)

    def desmarcar_todos(self):
        for iid in self.itens:
            self.itens[iid]["checked"] = False
            self.tree.set(iid, "sel", CHECK_OFF)

    def participantes_marcados(self):
        return [info["nome"] for info in self.itens.values() if info["checked"]]

    def participantes_nao_marcados(self):
        return [info["nome"] for info in self.itens.values() if not info["checked"]]

    # ---------- Cadastro ----------
    def add_membro(self):
        nome = self.nome_var.get().strip()
        secao = self.secao_var.get().strip()

        if not nome:
            messagebox.showerror("Erro", "Informe o nome.")
            return

        db_add(nome, secao)
        self.nome_var.set("")
        self.refresh_list()

    # ---------- Import/Export ----------
    def importar_membros(self):
        path = filedialog.askopenfilename(
            title="Importar membros",
            filetypes=[("CSV ou Excel", "*.csv *.xlsx *.xls"), ("CSV", "*.csv"), ("Excel", "*.xlsx *.xls")],
        )
        if not path:
            return

        try:
            rows = read_rows_from_file(path)
            if not rows:
                messagebox.showerror("Erro", "Arquivo vazio.")
                return

            adicionados = 0
            marcados = 0

            for r in rows:
                nome, secao = normalize_cols(r)
                if not nome:
                    continue

                found = db_find_by_name(nome)
                if not found:
                    db_add(nome, secao or "Outro")
                    adicionados += 1

            # Recarrega e marca os importados
            self.refresh_list()
            # marca os nomes do arquivo
            nomes_set = set()
            for r in rows:
                nome, _ = normalize_cols(r)
                if nome:
                    nomes_set.add(nome.lower())

            for iid, info in self.itens.items():
                if info["nome"].lower() in nomes_set:
                    info["checked"] = True
                    self.tree.set(iid, "sel", CHECK_ON)
                    marcados += 1

            messagebox.showinfo(
                "OK",
                f"Importação concluída.\nAdicionados ao cadastro: {adicionados}\nMarcados na lista: {marcados}",
            )
        except Exception as e:
            messagebox.showerror("Erro ao importar", str(e))

    def exportar_marcados(self):
        participantes = self.participantes_marcados()
        if not participantes:
            messagebox.showerror("Erro", "Não há membros marcados para exportar.")
            return

        caminho = filedialog.asksaveasfilename(
            title="Exportar marcados",
            defaultextension=".csv",
            initialfile="participantes_marcados.csv",
            filetypes=[("CSV", "*.csv")],
        )
        if not caminho:
            return

        rows = []
        # exporta nome e secao
        for info in self.itens.values():
            if info["checked"]:
                rows.append({"nome": info["nome"], "secao": info["secao"]})

        try:
            write_csv(caminho, rows)
            messagebox.showinfo("OK", f"Exportado em:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    # ---------- Pré-visualizar ----------
    def preview(self):
        participantes = self.participantes_marcados()
        if not participantes:
            messagebox.showerror("Erro", "Marque pelo menos 1 participante para pré-visualizar.")
            return

        atividade = self.atividade_var.get().strip()
        local = self.local_var.get().strip()
        dias = self.dias_var.get().strip()
        cidade = self.cidade_var.get().strip()
        linha_evento = self.linha_evento_var.get().strip()

        if not atividade or not local or not dias:
            messagebox.showerror("Erro", "Preencha Atividade, Local e Dias.")
            return

        try:
            dt = parse_data(self.data_var.get())
        except Exception:
            messagebox.showerror("Erro", "Data inválida. Use DD/MM/AAAA (ex: 19/02/2026).")
            return

        try:
            gerar_preview_pdf(participantes[0], atividade, local, dias, linha_evento, cidade, dt)
        except Exception as e:
            messagebox.showerror("Erro ao pré-visualizar", str(e))

    # ---------- Gerar ----------
    def gerar(self):
        participantes = self.participantes_marcados()
        if not participantes:
            messagebox.showerror("Erro", "Marque pelo menos 1 participante.")
            return

        atividade = self.atividade_var.get().strip()
        local = self.local_var.get().strip()
        dias = self.dias_var.get().strip()
        cidade = self.cidade_var.get().strip()
        linha_evento = self.linha_evento_var.get().strip()

        if not atividade or not local or not dias:
            messagebox.showerror("Erro", "Preencha Atividade, Local e Dias.")
            return

        # Data
        try:
            dt = parse_data(self.data_var.get())
        except Exception:
            messagebox.showerror("Erro", "Data inválida. Use DD/MM/AAAA (ex: 19/02/2026).")
            return

        # Aviso "não participaram"
        nao = self.participantes_nao_marcados()
        if nao:
            # mostra só os primeiros 15 pra não virar uma bíblia
            preview = "\n".join(nao[:15])
            if len(nao) > 15:
                preview += f"\n... (+{len(nao) - 15} nomes)"
            msg = (
                f"Você marcou {len(participantes)} participante(s).\n"
                f"Ficaram de fora {len(nao)}.\n\n"
                f"Fora da lista:\n{preview}\n\n"
                f"Deseja continuar e gerar apenas os marcados?"
            )
            if not messagebox.askyesno("Confirmar geração", msg):
                return

        # Salvar como (apenas atividade + data)
        data_br = dt.strftime("%d-%m-%Y")
        nome_padrao = f"{slug(atividade)}_{data_br}.pdf"

        caminho = filedialog.asksaveasfilename(
            title="Salvar certificados",
            defaultextension=".pdf",
            initialfile=nome_padrao,
            filetypes=[("Arquivo PDF", "*.pdf")],
        )
        if not caminho:
            return

        try:
            gerar_pdf(participantes, atividade, local, dias, linha_evento, cidade, dt, caminho)
            messagebox.showinfo("OK", f"PDF salvo em:\n{caminho}")
            save_hist_atividades(atividade)
            # atualiza combobox com histórico
            self.cb_atividade["values"] = load_hist_atividades()
        except Exception as e:
            messagebox.showerror("Erro ao gerar", str(e))


if __name__ == "__main__":
    App().mainloop()
